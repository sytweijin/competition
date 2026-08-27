"""Excel/CSV/ICS 导入导出服务。"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
import unicodedata
from datetime import date, datetime, timedelta, timezone
from xml.sax.saxutils import escape

from app.models.schemas import FullPlan, PlanOutput, ProjectModule, SubTask


def plan_fingerprint(raw: str) -> str:
    """计算方案内容的稳定指纹（sha1），用于并发冲突检测。

    剔除每次重算都会变化的 performance 与版本号字段，再做键排序序列化，
    保证"内容相同则指纹相同"。指纹只由后端计算，前端存储后端返回值，
    不存在前后端哈希算法不一致的问题。
    """
    data = json.loads(raw)
    data.pop("performance", None)
    data.pop("version", None)
    normalized = json.dumps(data, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()


def _as_date(value) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    text = str(value)[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _safe_text(value, default=""):
    return "" if value is None else str(value)


def _csv_section(
    writer, title: str, header: list[str], rows: list[list],
) -> None:
    """CSV 多区块结构：区块标题 + 空行 + 表头 + 数据行 + 空行。"""
    writer.writerow([title])
    writer.writerow([])
    writer.writerow(header)
    writer.writerows(rows)
    writer.writerow([])


def plan_to_csv(plan: FullPlan) -> str:
    """导出多区块 CSV：任务/成员/分工矩阵/时间线/参与清单/复盘，与 Excel 对齐。"""
    output = io.StringIO()
    writer = csv.writer(output)

    critical = {
        item.task_id: item.is_critical
        for item in (plan.timeline.tasks if plan.timeline else [])
    }
    task_rows = []
    for task in plan.plan.tasks:
        task_rows.append([
            task.id,
            task.name,
            task.module_id or "",
            task.estimated_hours,
            task.actual_hours if task.actual_hours is not None else "",
            task.assignee_id or "",
            ",".join(task.collaborator_ids or []),
            getattr(task, "suggested_people", None) or "",
            ",".join(task.required_skills or []),
            task.start_date.isoformat() if task.start_date else "",
            task.end_date.isoformat() if task.end_date else "",
            task.actual_end_date.isoformat() if task.actual_end_date else "",
            ",".join(task.dependencies or []),
            task.execution_stage or "",
            task.status.value if hasattr(task.status, "value") else task.status,
            "是" if critical.get(task.id) else "",
        ])
    _csv_section(
        writer, "任务",
        ["编号", "任务", "模块", "计划工时", "实际工时", "负责人",
         "协作者", "建议人数", "所需技能", "开始日期", "结束日期",
         "完成日期", "依赖", "阶段", "状态", "关键路径"],
        task_rows,
    )

    member_rows = [[
        m.name, m.role or "执行成员", m.manager or "",
        m.daily_available_hours, ",".join(m.skill_tags or []),
    ] for m in plan.input.members]
    _csv_section(
        writer, "成员",
        ["姓名", "角色", "上级", "每日可用工时", "技能"],
        member_rows,
    )

    matrix_rows = [[
        item.task_name, item.presenter or "", item.qa_primary or "",
        ",".join(item.qa_support or []), item.score,
    ] for item in plan.qa_matrix.assignments]
    _csv_section(
        writer, "分工矩阵",
        ["任务", "负责人", "主要协助", "辅助协助", "匹配度"],
        matrix_rows,
    )

    timeline_rows = [[
        item.name,
        item.start_date.isoformat() if item.start_date else "",
        item.end_date.isoformat() if item.end_date else "",
        "是" if item.is_critical else "",
        getattr(item, "float_days", 0),
    ] for item in (plan.timeline.tasks if plan.timeline else [])]
    _csv_section(
        writer, "时间线",
        ["任务", "开始", "结束", "关键", "浮动"],
        timeline_rows,
    )

    participant_rows = []
    for task in plan.plan.tasks:
        seen: set[str] = set()
        people: list[tuple[str, str, str]] = []
        if task.assignee_id:
            people.append((task.assignee_id, "负责人", "内部成员"))
        for collab in task.collaborator_ids or []:
            people.append((collab, "协作者", "内部成员"))
        for p in task.participants:
            people.append((
                p.name, p.role or "参与者",
                "志愿者 / 外部协作者" if p.is_volunteer else "内部成员",
            ))
        for v in plan.volunteer_pool or []:
            if v.task_id == task.id and v.status == "已确认":
                people.append((v.name, "志愿者", "志愿者 / 外部协作者"))
        for name, role, typ in people:
            if not name or name in seen:
                continue
            seen.add(name)
            participant_rows.append([
                f"{task.id} {task.name}", name, role, "", typ,
            ])
    _csv_section(
        writer, "参与清单",
        ["任务", "参与者", "角色", "投入工时", "类型"],
        participant_rows,
    )

    review_rows = []
    for task in plan.plan.tasks:
        dev = (
            round(task.actual_hours - task.estimated_hours, 2)
            if task.actual_hours is not None else ""
        )
        review_rows.append([
            task.name, task.estimated_hours,
            task.actual_hours if task.actual_hours is not None else "",
            dev,
            task.actual_end_date.isoformat() if task.actual_end_date else "",
        ])
    _csv_section(
        writer, "复盘",
        ["任务", "计划工时", "实际工时", "偏差", "实际完成"],
        review_rows,
    )
    # UTF-8 BOM：Windows Excel 直接打开中文 CSV 不乱码（ICS 已有 BOM）
    return "\ufeff" + output.getvalue()


def plan_to_ics(plan: FullPlan) -> str:
    timeline_map = {}
    if plan.timeline and plan.timeline.tasks:
        for item in plan.timeline.tasks:
            timeline_map[item.task_id] = item
    def esc_text(value: str) -> str:
        return value.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,")
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    cal_name = esc_text(plan.input.course.name or "协作分工方案")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Collaboration Planner//CN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{cal_name}",
        "X-WR-TIMEZONE:Asia/Shanghai",
    ]
    for task in plan.plan.tasks:
        start = task.start_date
        end = task.end_date
        tl = timeline_map.get(task.id)
        if start is None and tl is not None:
            start = _as_date(getattr(tl, "start_date", None))
        if end is None and tl is not None:
            end = _as_date(getattr(tl, "end_date", None))
        if start is None or end is None:
            continue
        uid = f"{task.id}@collaboration-planner"
        summary = esc_text(task.name)
        description = (
            f"{task.estimated_hours}h · {task.assignee_id or '未分配'}"
        )
        description = esc_text(description)
        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        lines.append(f"DTSTAMP:{stamp}")
        lines.append(f"DTSTART;VALUE=DATE:{start.strftime('%Y%m%d')}")
        lines.append(f"DTEND;VALUE=DATE:{(end + timedelta(days=1)).strftime('%Y%m%d')}")
        lines.append(f"SUMMARY:{summary}")
        lines.append(f"DESCRIPTION:{description}")
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    raw = "\r\n".join(lines) + "\r\n"
    # RFC 5545 折行：超过 75 字节的行用 CRLF+空格续行；加 UTF-8 BOM，
    # 保证 Windows/Outlook 能识别含中文的 ICS。
    def fold(line: str) -> list[str]:
        if len(line.encode("utf-8")) <= 75:
            return [line]
        out: list[str] = []
        cur = ""
        cur_bytes = 0
        for ch in line:
            ch_bytes = len(ch.encode("utf-8"))
            limit = 74 if out else 75
            if cur_bytes + ch_bytes > limit:
                out.append(cur)
                cur = " "
                cur_bytes = 1
            cur += ch
            cur_bytes += ch_bytes
        if cur:
            out.append(cur)
        return out

    folded: list[str] = []
    for line in raw.split("\r\n"):
        folded.extend(fold(line))
    return "\ufeff" + "\r\n".join(folded) + "\r\n"


def _xlsx_bytes(sheets: list[tuple[str, list[list]]]) -> bytes:
    """极简 XLSX 生成器，不依赖 openpyxl。"""
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/content-types"
    ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_rel_ws = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    content = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<Types xmlns="{ns_pkg}">',
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
    ]
    for i in range(1, len(sheets) + 1):
        content.append(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    content.append("</Types>")

    workbook = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<workbook xmlns="{ns_main}" xmlns:r="{ns_rel_ws}"><sheets>',
    ]
    for i, (name, _) in enumerate(sheets, start=1):
        workbook.append(
            f'<sheet name="{escape(name)}" sheetId="{i}" r:id="rId{i}"/>'
        )
    workbook.append("</sheets></workbook>")

    workbook_rels = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<Relationships xmlns="{ns_rel}">',
    ]
    for i in range(1, len(sheets) + 1):
        workbook_rels.append(
            f'<Relationship Id="rId{i}" '
            f'Type="{ns_rel_ws}/worksheet" Target="worksheets/sheet{i}.xml"/>'
        )
    workbook_rels.append("</Relationships>")

    sheet_xmls = []
    for _, rows in sheets:
        xml_rows = []
        for row_idx, row in enumerate(rows, start=1):
            cells = []
            for col_idx, value in enumerate(row, start=1):
                ref = f"{chr(64 + col_idx)}{row_idx}"
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cells.append(f'<c r="{ref}"><v>{value}</v></c>')
                else:
                    text = escape("" if value is None else str(value))
                    cells.append(
                        f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'
                    )
            xml_rows.append("<row>" + "".join(cells) + "</row>")
        sheet_xmls.append(
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<worksheet xmlns="{ns_main}"><sheetData>'
            + "".join(xml_rows)
            + "</sheetData></worksheet>"
        )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "\n".join(content))
        zf.writestr("_rels/.rels", (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{ns_rel}">'
            f'<Relationship Id="rId1" Type="{ns_rel}/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>"
        ))
        zf.writestr("xl/workbook.xml", "\n".join(workbook))
        zf.writestr("xl/_rels/workbook.xml.rels", "\n".join(workbook_rels))
        for i, sheet_xml in enumerate(sheet_xmls, start=1):
            zf.writestr(f"xl/worksheets/sheet{i}.xml", sheet_xml)
    return buffer.getvalue()


def plan_to_excel(plan: FullPlan) -> bytes:
    """导出 Excel 工作簿（任务/成员/分工/时间线/参与清单/复盘）。"""
    sheets: list[tuple[str, list[list]]] = []

    task_rows = [[
        "编号", "任务", "模块", "计划工时", "实际工时", "负责人",
        "开始日期", "结束日期", "依赖", "阶段", "状态",
    ]]
    for task in plan.plan.tasks:
        task_rows.append([
            task.id, task.name, task.module_id or "", task.estimated_hours,
            task.actual_hours if task.actual_hours is not None else "",
            task.assignee_id or "",
            task.start_date.isoformat() if task.start_date else "",
            task.end_date.isoformat() if task.end_date else "",
            ",".join(task.dependencies or []),
            task.execution_stage or "",
            task.status.value if hasattr(task.status, "value") else task.status,
        ])
    sheets.append(("任务", task_rows))

    member_rows = [["姓名", "角色", "上级", "每日可用工时", "技能"]]
    for member in plan.input.members:
        member_rows.append([
            member.name, member.role or "执行成员", member.manager or "",
            member.daily_available_hours, ",".join(member.skill_tags or []),
        ])
    sheets.append(("成员", member_rows))

    matrix_rows = [["任务", "负责人", "主要协助", "辅助协助", "匹配度"]]
    for item in plan.qa_matrix.assignments:
        matrix_rows.append([
            item.task_name, item.presenter, item.qa_primary,
            ",".join(item.qa_support or []),
            item.score,
        ])
    sheets.append(("分工矩阵", matrix_rows))

    timeline_rows = [["任务", "开始", "结束", "关键", "浮动"]]
    for item in plan.timeline.tasks:
        timeline_rows.append([
            item.name, item.start_date, item.end_date,
            "是" if item.is_critical else "",
            getattr(item, "float_days", 0),
        ])
    sheets.append(("时间线", timeline_rows))

    participant_rows = [["任务", "参与者", "角色", "投入工时", "类型"]]
    for task in plan.plan.tasks:
        seen: set[str] = set()
        people: list[tuple[str, str, str]] = []
        if task.assignee_id:
            people.append((task.assignee_id, "负责人", "内部成员"))
        for collab in task.collaborator_ids or []:
            people.append((collab, "协作者", "内部成员"))
        for p in task.participants:
            people.append((
                p.name, p.role or "参与者",
                "志愿者 / 外部协作者" if p.is_volunteer else "内部成员",
            ))
        for v in plan.volunteer_pool or []:
            if v.task_id == task.id and v.status == "已确认":
                people.append((v.name, "志愿者", "志愿者 / 外部协作者"))
        for name, role, typ in people:
            if not name or name in seen:
                continue
            seen.add(name)
            participant_rows.append([
                f"{task.id} {task.name}", name, role, "", typ,
            ])
    sheets.append(("参与清单", participant_rows))

    review_rows = [["任务", "计划工时", "实际工时", "偏差", "实际完成"]]
    for task in plan.plan.tasks:
        dev = (
            round(task.actual_hours - task.estimated_hours, 2)
            if task.actual_hours is not None else ""
        )
        review_rows.append([
            task.name, task.estimated_hours, task.actual_hours, dev,
            task.actual_end_date.isoformat() if task.actual_end_date else "",
        ])
    sheets.append(("复盘", review_rows))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法生成标准 Excel 工作簿") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(title=sheet_name[:31])
        for row in rows:
            worksheet.append(row)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index, column in enumerate(worksheet.columns, start=1):
            width = max(sum(
                2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
                for char in str(cell.value or "")
            ) for cell in column)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(width + 3, 10), 42
            )
        worksheet.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_task_file(content: bytes, filename: str, project_mode: str = "small_group") -> PlanOutput:
    """从 CSV/Excel 第一张表导入任务草稿。"""
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("openpyxl 未安装，无法导入 Excel") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    elif lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        raise ValueError("仅支持 .csv / .xlsx 文件导入")

    if not rows:
        raise ValueError("文件为空")
    headers = [str(c or "").strip().lower() for c in rows[0]]
    data_rows = rows[1:]

    def col(*names):
        for idx, header in enumerate(headers):
            if any(name in header for name in names):
                return idx
        return -1

    id_idx = col("编号", "id")
    name_idx = col("任务", "名称", "name")
    module_idx = col("模块", "module")
    hours_idx = col("计划工时", "工时", "hours")
    actual_idx = col("实际工时", "actual")
    owner_idx = col("负责人", "assignee", "owner")
    start_idx = col("开始日期", "start")
    end_idx = col("结束日期", "end", "截止")
    dep_idx = col("依赖", "dependencies")
    stage_idx = col("阶段", "stage")
    skill_idx = col("技能", "skills")

    if name_idx < 0:
        raise ValueError("缺少“任务/名称”列")

    tasks: list[SubTask] = []
    module_names: list[str] = []
    for idx, row in enumerate(data_rows, start=1):
        def cell(ci):
            if ci < 0 or ci >= len(row):
                return ""
            value = row[ci]
            return "" if value is None else str(value).strip()

        name = cell(name_idx)
        if not name:
            continue
        module_name = cell(module_idx)
        if module_name and module_name not in module_names:
            module_names.append(module_name)
        try:
            hours = max(0.5, float(cell(hours_idx) or 2))
        except ValueError:
            hours = 2.0
        actual = None
        if cell(actual_idx):
            try:
                actual = max(0.0, float(cell(actual_idx)))
            except ValueError:
                actual = None
        deps = [d.strip() for d in re.split(r"[,，;；]", cell(dep_idx)) if d.strip()]
        task_id = cell(id_idx) or f"T{idx}"
        tasks.append(SubTask(
            id=task_id,
            module_id=f"M{module_names.index(module_name) + 1}" if module_name else None,
            name=name,
            description="",
            estimated_hours=round(hours, 2),
            actual_hours=actual,
            assignee_id=cell(owner_idx) or None,
            start_date=_as_date(cell(start_idx)),
            end_date=_as_date(cell(end_idx)),
            dependencies=deps,
            execution_stage=cell(stage_idx) or "执行",
            required_skills=[
                s.strip() for s in re.split(r"[,，;；]", cell(skill_idx))
                if s.strip()
            ],
        ))

    if not tasks:
        raise ValueError("没有可导入的任务行")
    modules = [
        ProjectModule(id=f"M{i + 1}", name=name, order=i + 1)
        for i, name in enumerate(module_names)
    ]
    if project_mode != "large_project":
        modules = []
    return PlanOutput(tasks=tasks, modules=modules, summary=f"导入 {len(tasks)} 项任务")
