"""Excel/CSV/ICS import/export tests."""

from datetime import date

import pytest

from app.models.schemas import (
    AssignmentInput, CourseInfo, FullPlan, PlanOutput, QAOutput, ReportOutput,
    SubTask, TeamMember, TimelineOutput,
)
from app.services.plan_io import (
    parse_task_file, plan_to_csv, plan_to_excel, plan_to_ics,
)


def _plan():
    return FullPlan(
        input=AssignmentInput(
            course=CourseInfo(name="测试", description=""),
            members=[TeamMember(name="小文", role="执行成员")],
            deadline=date(2026, 8, 20),
        ),
        plan=PlanOutput(
            tasks=[
                SubTask(
                    id="T1", name="调研", estimated_hours=4,
                    assignee_id="小文",
                    start_date=date(2026, 8, 5),
                    end_date=date(2026, 8, 6),
                ),
            ],
            summary="测试",
        ),
        timeline=TimelineOutput(tasks=[], critical_path=[], total_days=0),
        qa_matrix=QAOutput(assignments=[]),
        report=ReportOutput(summary=""),
    )


def test_plan_to_csv_and_ics():
    plan = _plan()
    csv_text = plan_to_csv(plan)
    assert csv_text.startswith("\ufeff")  # Excel 中文兼容
    assert "编号" in csv_text
    assert "T1" in csv_text
    assert "调研" in csv_text
    ics_text = plan_to_ics(plan)
    assert "BEGIN:VEVENT" in ics_text
    assert "SUMMARY:调研" in ics_text
    assert "DTEND;VALUE=DATE:20260807" in ics_text


def test_plan_to_ics_bom_method_and_folding():
    """ICS 带 UTF-8 BOM、METHOD:PUBLISH，长行按 RFC5545 折行。"""
    plan = _plan()
    plan.plan.tasks[0].name = "很长" * 40  # 触发折行
    text = plan_to_ics(plan)
    assert text.startswith("\ufeff")
    assert "METHOD:PUBLISH" in text
    lines = text.lstrip("\ufeff").split("\r\n")
    for line in lines:
        assert len(line.encode("utf-8")) <= 75
    assert any(line.startswith(" ") for line in lines)  # 续行以空格开头


def test_plan_to_excel_participants_and_review():
    """参与清单由负责人/协作者/志愿者推导；复盘列出全部任务（含无实际工时）。"""
    pytest.importorskip("openpyxl")
    import io
    from openpyxl import load_workbook

    plan = _plan()
    plan.plan.tasks[0].collaborator_ids = ["小红"]
    data = plan_to_excel(plan)
    wb = load_workbook(io.BytesIO(data))
    parts = list(wb["参与清单"].iter_rows(values_only=True))
    assert len(parts) == 3  # 表头 + 负责人 + 协作者
    assert parts[1][1] == "小文"
    assert parts[2][1] == "小红"
    rev = list(wb["复盘"].iter_rows(values_only=True))
    assert len(rev) == 2  # 表头 + 任务（无实际工时也列出）


def test_plan_to_csv_extra_columns():
    """CSV 多区块：任务/成员/分工/时间线/参与清单/复盘，列补全。"""
    plan = _plan()
    plan.plan.tasks[0].collaborator_ids = ["小红"]
    plan.plan.tasks[0].required_skills = ["调研"]
    csv_text = plan_to_csv(plan)
    assert "协作者" in csv_text
    assert "关键路径" in csv_text
    assert "所需技能" in csv_text
    assert "小红" in csv_text
    assert "调研" in csv_text
    # 区块标题与 Excel 的 sheet 对齐
    for section in ("任务", "成员", "分工矩阵", "时间线", "参与清单", "复盘"):
        assert section + "\r\n" in csv_text
    # 复盘区块即使无实际工时也会列出任务
    assert "小文" in csv_text.split("参与清单")[1]


def test_parse_task_file_csv():
    content = (
        "编号,任务,模块,计划工时,负责人,开始日期,结束日期\n"
        "T1,调研,M1,4,小文,2026-08-05,2026-08-06\n"
    ).encode("utf-8")
    plan = parse_task_file(content, "tasks.csv", "large_project")
    assert len(plan.tasks) == 1
    assert plan.tasks[0].id == "T1"
    assert plan.tasks[0].module_id == "M1"
    assert plan.tasks[0].estimated_hours == 4


def test_plan_to_excel():
    pytest.importorskip("openpyxl")
    data = plan_to_excel(_plan())
    assert data[:2] == b"PK"
    from openpyxl import load_workbook

    import io

    wb = load_workbook(io.BytesIO(data))
    assert "任务" in wb.sheetnames
    assert "成员" in wb.sheetnames
    assert "复盘" in wb.sheetnames
    assert wb["任务"].freeze_panes == "A2"
    assert wb["任务"].auto_filter.ref == wb["任务"].dimensions
    assert wb["任务"]["A1"].fill.fill_type == "solid"
